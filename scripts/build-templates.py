#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generátor šablon ke stažení pro /sablony.

Každá šablona vychází z konkrétního návodu na webu a je to skutečný funkční
soubor — .xlsx se živými vzorci (SUMIFS, COUNTIFS, EDATE), .csv připravené
k importu. Ke každé šabloně vzniká i anglická mutace `<nazev>-en.<pripona>`
s anglickými názvy listů a sloupců.

Spuštění:  python3 scripts/build-templates.py
Závislost: openpyxl  (pip install openpyxl)

Žádná skutečná data — všechny ukázkové řádky jsou vymyšlené.
"""

from __future__ import annotations

import csv
import datetime as dt
import os
import pathlib

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "public" / "sablony"

# ---------------------------------------------------------------- design
# Papír, inkoust, akcent jen na signál (záporné cash flow). Stejné hodnoty
# jako v app/globals.css, aby šablony vypadaly jako zbytek webu.
INK = "FF141414"
MUTED = "FF494B4C"
FAINT = "FF6C6E6F"
ACCENT = "FFD63210"
SURFACE = "FFF7F7F5"

F_TITLE = Font(name="Calibri", size=15, bold=True, color=INK)
F_NOTE = Font(name="Calibri", size=10, italic=True, color=FAINT)
F_HEAD = Font(name="Calibri", size=11, bold=True, color=INK)
F_SECTION = Font(name="Calibri", size=12, bold=True, color=INK)
F_BODY = Font(name="Calibri", size=11, color=INK)
F_COMMENT = Font(name="Calibri", size=10, color=MUTED)
F_TOTAL = Font(name="Calibri", size=11, bold=True, color=INK)
F_RESULT = Font(name="Calibri", size=13, bold=True, color=INK)

FILL_HEAD = PatternFill("solid", fgColor=SURFACE)
BORDER_HEAD = Border(bottom=Side(style="medium", color=INK))
BORDER_TOP = Border(top=Side(style="thin", color=INK))
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")

MONEY = '#,##0" Kč"'
MONEY_EN = '#,##0" CZK"'
MONEY2 = '#,##0.00" Kč"'
MONEY2_EN = '#,##0.00" CZK"'
PCT = "0 %"
DATEF = "yyyy-mm-dd"
MONTHF = "yyyy-mm"

ROWS_TX = 400          # dokud vzorce v listu Transakce dosahují
ROWS_ING = 200         # dokud dosahují vzorce v surovinách
MONTHS = 12


def q(sheet: str) -> str:
    """Odkaz na list ve vzorci. Apostrofy vždycky — název s mezerou i bez."""
    return "'" + sheet.replace("'", "''") + "'!"


def head(ws, row: int, labels: list[str], widths: list[int] | None = None) -> None:
    for i, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER_HEAD
        c.alignment = WRAP
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def note(ws, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = F_NOTE
    ws[cell].alignment = TOP


# ================================================================ 1. rozpočet
# Zdroj: content/tipy/osobni-rozpocet-z-bankovniho-vypisu.mdx

BUDGET = {
    "cs": {
        "file": "osobni-rozpocet.xlsx",
        "sheets": ["Transakce", "Kategorie", "Přehled po měsících", "Cash flow"],
        "money": MONEY,
        "money2": MONEY2,
        "tx_head": ["Datum", "Popis", "Částka", "Kategorie", "Typ platby", "Měsíc"],
        "tx_note": (
            "Sem vložte anonymizovaný export z banky. Výdaje jako záporná čísla, "
            "příjmy jako kladná. Sloupec Měsíc se počítá sám, nepřepisujte ho. "
            "Vlastní převody mezi svými účty dejte do kategorie „Vlastní převod“ — "
            "do rozpočtu se nezapočítají."
        ),
        "cat_head": ["Kategorie", "Typ", "Poznámka"],
        "pay_head": ["Typ platby"],
        "cat_note": (
            "Kategorie si upravte podle sebe, ale držte se deseti až patnácti — "
            "u víc kategorií přestanete vlastní tabulce rozumět. Názvy se propisují "
            "do rozevíracího seznamu v Transakcích i do přehledu."
        ),
        "types": {"exp": "výdaj", "inc": "příjem", "skip": "mimo rozpočet"},
        "payments": ["karta", "převod", "inkaso", "výběr", "hotovost", "poplatek"],
        "categories": [
            ("Bydlení", "exp", "Nájem nebo hypotéka, energie, internet, poplatky."),
            ("Jídlo – nákupy", "exp", "Potraviny z obchodu. Zvlášť od restaurací a rozvozu."),
            ("Jídlo – restaurace", "exp", "Obědy a večeře venku."),
            ("Jídlo – rozvoz", "exp", "Rozvoz jídla. Nejčastější překvapení celé analýzy."),
            ("Doprava", "exp", "MHD, palivo, servis, taxi, parkování."),
            ("Předplatná a služby", "exp", "Opakované platby za digitální i jiné služby."),
            ("Zdraví", "exp", "Léky, lékaři, terapie, pojištění zdraví."),
            ("Děti", "exp", "Kroužky, školky, výbava, tábory."),
            ("Oblečení", "exp", "Oblečení a obuv."),
            ("Zábava a koníčky", "exp", "Kultura, sport, vybavení na koníčky."),
            ("Dárky", "exp", "Dárky a příspěvky."),
            ("Jednorázové velké výdaje", "exp", "Pračka, kolo, dovolená, kauce. Bez oddělení rozhodí průměr."),
            ("Splátky a spoření", "exp", "Splátky úvěrů a odkládané peníze."),
            ("Ostatní", "exp", "Co se nikam nevešlo."),
            ("Nezařazeno", "exp", "Fronta k ručnímu dotřídění. Nikdy nehádejte."),
            ("Příjem – pravidelný", "inc", "Mzda, pravidelné faktury."),
            ("Příjem – nepravidelný", "inc", "Bonusy, prodeje, jednorázové zakázky."),
            ("Vlastní převod", "skip", "Přesun mezi vlastními účty. Není příjem ani výdaj."),
        ],
        "ov_title": "Přehled po měsících",
        "ov_note": (
            "Čísla se počítají sama z listu Transakce funkcí SUMIFS. Nic sem nepište "
            "ručně — stačí změnit první měsíc níž a celá tabulka se posune."
        ),
        "ov_start": "První měsíc přehledu",
        "ov_head": ["Kategorie", "Typ", "Celkem"],
        "ov_rows": {
            "exp": "Výdaje celkem",
            "inc": "Příjmy celkem",
            "cf": "Cash flow",
            "cfx": "Cash flow bez jednorázových výdajů",
        },
        "cf_head": [
            "Měsíc", "Příjmy", "Výdaje", "Cash flow",
            "Cash flow bez jednorázových", "Poznámka",
        ],
        "cf_note": (
            "Cash flow je přítok minus odtok, ne zůstatek na účtu. Záporný měsíc "
            "znamená, že vás platí to, co jste našetřili dřív. Sledujte řadu, "
            "ne jedno číslo."
        ),
        "cf_neg": "záporný měsíc",
        "cf_sum": [
            ("Součet za období", "SUM"),
            ("Průměr měsíčně", "AVERAGE"),
            ("Medián měsíčně", "MEDIAN"),
        ],
        "cf_negcount": "Počet záporných měsíců",
        "cf_medhint": "Medián říká víc než průměr — jeden nákup pračky průměr zvedne, medián ne.",
    },
    "en": {
        "file": "osobni-rozpocet-en.xlsx",
        "sheets": ["Transactions", "Categories", "Monthly overview", "Cash flow"],
        "money": MONEY_EN,
        "money2": MONEY2_EN,
        "tx_head": ["Date", "Description", "Amount", "Category", "Payment type", "Month"],
        "tx_note": (
            "Paste your anonymised bank export here. Expenses as negative numbers, "
            "income as positive. The Month column calculates itself — do not overwrite it. "
            "Transfers between your own accounts go to the “Own transfer” category and "
            "stay out of the budget."
        ),
        "cat_head": ["Category", "Type", "Note"],
        "pay_head": ["Payment type"],
        "cat_note": (
            "Adjust the categories to fit you, but keep it to ten or fifteen — beyond that "
            "you stop understanding your own table. The names feed the drop-down in "
            "Transactions and the monthly overview."
        ),
        "types": {"exp": "expense", "inc": "income", "skip": "outside budget"},
        "payments": ["card", "transfer", "direct debit", "withdrawal", "cash", "fee"],
        "categories": [
            ("Housing", "exp", "Rent or mortgage, utilities, internet, fees."),
            ("Food – groceries", "exp", "Shopping. Kept apart from eating out and delivery."),
            ("Food – eating out", "exp", "Lunches and dinners out."),
            ("Food – delivery", "exp", "Food delivery. The most common surprise of the whole analysis."),
            ("Transport", "exp", "Public transport, fuel, service, taxi, parking."),
            ("Subscriptions and services", "exp", "Recurring payments, digital and otherwise."),
            ("Health", "exp", "Medicines, doctors, therapy, health insurance."),
            ("Children", "exp", "Clubs, nursery, gear, camps."),
            ("Clothes", "exp", "Clothing and shoes."),
            ("Fun and hobbies", "exp", "Culture, sport, hobby gear."),
            ("Gifts", "exp", "Presents and contributions."),
            ("One-off big spends", "exp", "Washing machine, bike, holiday, deposit. Wrecks the average if mixed in."),
            ("Repayments and savings", "exp", "Loan repayments and money set aside."),
            ("Other", "exp", "Whatever did not fit anywhere."),
            ("Unclassified", "exp", "The manual queue. Never guess."),
            ("Income – regular", "inc", "Salary, regular invoices."),
            ("Income – irregular", "inc", "Bonuses, sales, one-off jobs."),
            ("Own transfer", "skip", "A move between your own accounts. Neither income nor expense."),
        ],
        "ov_title": "Monthly overview",
        "ov_note": (
            "The numbers come from the Transactions sheet through SUMIFS. Do not type "
            "anything in here — change the first month below and the whole table shifts."
        ),
        "ov_start": "First month of the overview",
        "ov_head": ["Category", "Type", "Total"],
        "ov_rows": {
            "exp": "Total expenses",
            "inc": "Total income",
            "cf": "Cash flow",
            "cfx": "Cash flow without one-off spends",
        },
        "cf_head": [
            "Month", "Income", "Expenses", "Cash flow",
            "Cash flow without one-offs", "Note",
        ],
        "cf_note": (
            "Cash flow is money in minus money out, not the balance on your account. "
            "A negative month means you are being paid by what you saved earlier. "
            "Watch the series, not a single number."
        ),
        "cf_neg": "negative month",
        "cf_sum": [
            ("Total for the period", "SUM"),
            ("Monthly average", "AVERAGE"),
            ("Monthly median", "MEDIAN"),
        ],
        "cf_negcount": "Negative months",
        "cf_medhint": "The median says more than the average — one washing machine lifts the average, not the median.",
    },
}

# Vymyšlené transakce: tři měsíce, jeden záporný. Index kategorie odkazuje
# do seznamu výš, aby obě jazykové mutace seděly na stejná data.
BUDGET_TX = [
    ("2026-01-05", "Nájem byt", -14500, 0, 1),
    ("2026-01-05", "Elektřina záloha", -2100, 0, 2),
    ("2026-01-08", "Supermarket týdenní nákup", -1840, 1, 0),
    ("2026-01-09", "Streamovací služba", -279, 5, 2),
    ("2026-01-12", "Rozvoz jídla", -520, 3, 0),
    ("2026-01-15", "Faktura klient A", 38000, 15, 1),
    ("2026-01-16", "MHD kupón čtvrtletní", -1280, 4, 1),
    ("2026-01-18", "Supermarket týdenní nákup", -2110, 1, 0),
    ("2026-01-22", "Lékárna", -430, 6, 0),
    ("2026-01-25", "Převod na spořicí účet", -5000, 17, 1),
    ("2026-01-27", "Restaurace oběd", -390, 2, 0),
    ("2026-02-05", "Nájem byt", -14500, 0, 1),
    ("2026-02-06", "Supermarket týdenní nákup", -1990, 1, 0),
    ("2026-02-09", "Streamovací služba", -279, 5, 2),
    ("2026-02-11", "Pračka náhrada za rozbitou", -12490, 11, 0),
    ("2026-02-14", "Rozvoz jídla", -680, 3, 0),
    ("2026-02-15", "Faktura klient A", 32000, 15, 1),
    ("2026-02-19", "Kroužek plavání", -1200, 7, 1),
    ("2026-02-21", "Supermarket týdenní nákup", -2260, 1, 0),
    ("2026-02-24", "Bunda", -1890, 8, 0),
    ("2026-02-27", "Výběr z bankomatu", -2000, 14, 3),
    ("2026-03-05", "Nájem byt", -14500, 0, 1),
    ("2026-03-07", "Supermarket týdenní nákup", -1770, 1, 0),
    ("2026-03-09", "Streamovací služba", -279, 5, 2),
    ("2026-03-12", "Antivir roční platba", -1490, 5, 1),
    ("2026-03-15", "Faktura klient A", 41000, 15, 1),
    ("2026-03-16", "Faktura klient B jednorázová", 9500, 16, 1),
    ("2026-03-18", "Kino", -520, 9, 0),
    ("2026-03-20", "Supermarket týdenní nákup", -2050, 1, 0),
    ("2026-03-24", "Dárek k narozeninám", -900, 10, 0),
    ("2026-03-28", "Servis kola", -1350, 4, 0),
]

BUDGET_TX_EN = {
    "Nájem byt": "Rent",
    "Elektřina záloha": "Electricity advance",
    "Supermarket týdenní nákup": "Supermarket weekly shop",
    "Streamovací služba": "Streaming service",
    "Rozvoz jídla": "Food delivery",
    "Faktura klient A": "Invoice client A",
    "MHD kupón čtvrtletní": "Transit pass quarterly",
    "Lékárna": "Pharmacy",
    "Převod na spořicí účet": "Transfer to savings account",
    "Restaurace oběd": "Restaurant lunch",
    "Pračka náhrada za rozbitou": "Washing machine replacement",
    "Kroužek plavání": "Swimming club",
    "Bunda": "Jacket",
    "Výběr z bankomatu": "ATM withdrawal",
    "Antivir roční platba": "Antivirus annual payment",
    "Faktura klient B jednorázová": "Invoice client B one-off",
    "Kino": "Cinema",
    "Dárek k narozeninám": "Birthday present",
    "Servis kola": "Bike service",
}


def build_budget(loc: str) -> pathlib.Path:
    t = BUDGET[loc]
    s_tx, s_cat, s_ov, s_cf = t["sheets"]
    cats = t["categories"]
    n_cat = len(cats)
    exp_idx = [i for i, c in enumerate(cats) if c[1] == "exp"]
    inc_idx = [i for i, c in enumerate(cats) if c[1] == "inc"]
    oneoff_idx = next(i for i, c in enumerate(cats) if "one-off" in c[0].lower() or "Jednorázové" in c[0])

    wb = Workbook()

    # ---- Kategorie (číselník) --------------------------------------------
    ws = wb.active
    ws.title = s_cat
    head(ws, 1, t["cat_head"], [30, 16, 60])
    for i, (name, kind, desc) in enumerate(cats):
        r = i + 2
        ws.cell(row=r, column=1, value=name).font = F_BODY
        ws.cell(row=r, column=2, value=t["types"][kind]).font = F_BODY
        c = ws.cell(row=r, column=3, value=desc)
        c.font = F_COMMENT
        c.alignment = WRAP
    ws.cell(row=1, column=5, value=t["pay_head"][0]).font = F_HEAD
    ws.cell(row=1, column=5).fill = FILL_HEAD
    ws.cell(row=1, column=5).border = BORDER_HEAD
    ws.column_dimensions["E"].width = 18
    for i, p in enumerate(t["payments"]):
        ws.cell(row=i + 2, column=5, value=p).font = F_BODY
    ws.column_dimensions["G"].width = 46
    note(ws, "G1", t["cat_note"])
    ws.freeze_panes = "A2"
    cat_last = n_cat + 1
    pay_last = len(t["payments"]) + 1

    # ---- Transakce --------------------------------------------------------
    ws = wb.create_sheet(s_tx)
    head(ws, 1, t["tx_head"], [13, 34, 15, 26, 15, 11])
    for i, (date, desc, amount, cat_i, pay_i) in enumerate(BUDGET_TX):
        r = i + 2
        c = ws.cell(row=r, column=1, value=dt.date.fromisoformat(date))
        c.number_format = DATEF
        c.font = F_BODY
        label = BUDGET_TX_EN[desc] if loc == "en" else desc
        ws.cell(row=r, column=2, value=label).font = F_BODY
        c = ws.cell(row=r, column=3, value=amount)
        c.number_format = t["money2"]
        c.font = F_BODY
        ws.cell(row=r, column=4, value=cats[cat_i][0]).font = F_BODY
        ws.cell(row=r, column=5, value=t["payments"][pay_i]).font = F_BODY
    for r in range(2, ROWS_TX + 1):
        c = ws.cell(row=r, column=6, value=f'=IF($A{r}="","",TEXT($A{r},"yyyy-mm"))')
        c.font = F_COMMENT
        if r > len(BUDGET_TX) + 1:
            ws.cell(row=r, column=1).number_format = DATEF
            ws.cell(row=r, column=3).number_format = t["money2"]

    dv_cat = DataValidation(
        type="list", formula1=f"={q(s_cat)}$A$2:$A${cat_last}", allow_blank=True
    )
    dv_pay = DataValidation(
        type="list", formula1=f"={q(s_cat)}$E$2:$E${pay_last}", allow_blank=True
    )
    ws.add_data_validation(dv_cat)
    ws.add_data_validation(dv_pay)
    dv_cat.add(f"D2:D{ROWS_TX}")
    dv_pay.add(f"E2:E{ROWS_TX}")

    ws.column_dimensions["H"].width = 52
    note(ws, "H1", t["tx_note"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ROWS_TX}"

    # ---- Přehled po měsících ---------------------------------------------
    ws = wb.create_sheet(s_ov)
    ws["A1"] = t["ov_title"]
    ws["A1"].font = F_TITLE
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    note(ws, "A2", t["ov_note"])
    ws["A3"] = t["ov_start"]
    ws["A3"].font = F_HEAD
    ws["B3"] = dt.date(2026, 1, 1)
    ws["B3"].number_format = MONTHF
    ws["B3"].font = F_BODY

    HR = 5                      # řádek hlavičky
    first = HR + 1              # první kategorie
    ws.cell(row=HR, column=1, value=t["ov_head"][0])
    ws.cell(row=HR, column=2, value=t["ov_head"][1])
    for m in range(MONTHS):
        col = 3 + m
        f = f'=TEXT($B$3,"yyyy-mm")' if m == 0 else f'=TEXT(EDATE($B$3,{m}),"yyyy-mm")'
        ws.cell(row=HR, column=col, value=f)
        ws.column_dimensions[get_column_letter(col)].width = 12
    total_col = 3 + MONTHS
    ws.cell(row=HR, column=total_col, value=t["ov_head"][2])
    ws.column_dimensions[get_column_letter(total_col)].width = 14
    for col in range(1, total_col + 1):
        c = ws.cell(row=HR, column=col)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER_HEAD

    for i in range(n_cat):
        r = first + i
        ws.cell(row=r, column=1, value=f"={q(s_cat)}A{i + 2}").font = F_BODY
        ws.cell(row=r, column=2, value=f"={q(s_cat)}B{i + 2}").font = F_COMMENT
        for m in range(MONTHS):
            col = get_column_letter(3 + m)
            f = (
                f"=SUMIFS({q(s_tx)}$C$2:$C${ROWS_TX},"
                f"{q(s_tx)}$D$2:$D${ROWS_TX},$A{r},"
                f"{q(s_tx)}$F$2:$F${ROWS_TX},{col}${HR})"
            )
            c = ws.cell(row=r, column=3 + m, value=f)
            c.number_format = t["money"]
            c.font = F_BODY
        c = ws.cell(row=r, column=total_col, value=f"=SUM(C{r}:{get_column_letter(2 + MONTHS)}{r})")
        c.number_format = t["money"]
        c.font = F_TOTAL

    exp_from, exp_to = first + exp_idx[0], first + exp_idx[-1]
    inc_from, inc_to = first + inc_idx[0], first + inc_idx[-1]
    oneoff_row = first + oneoff_idx

    r_exp = first + n_cat + 1
    r_inc = r_exp + 1
    r_cf = r_exp + 2
    r_cfx = r_exp + 3
    labels = [t["ov_rows"]["exp"], t["ov_rows"]["inc"], t["ov_rows"]["cf"], t["ov_rows"]["cfx"]]
    for i, label in enumerate(labels):
        r = r_exp + i
        c = ws.cell(row=r, column=1, value=label)
        c.font = F_TOTAL
        c.border = BORDER_TOP if i == 0 else Border()
        for m in range(MONTHS + 1):
            col = get_column_letter(3 + m)
            if i == 0:
                f = f"=SUM({col}{exp_from}:{col}{exp_to})"
            elif i == 1:
                f = f"=SUM({col}{inc_from}:{col}{inc_to})"
            elif i == 2:
                f = f"={col}{r_inc}+{col}{r_exp}"
            else:
                f = f"={col}{r_cf}-{col}{oneoff_row}"
            c = ws.cell(row=r, column=3 + m, value=f)
            c.number_format = t["money"]
            c.font = F_TOTAL
            if i == 0:
                c.border = BORDER_TOP

    last_col = get_column_letter(total_col)
    ws.conditional_formatting.add(
        f"C{r_cf}:{last_col}{r_cfx}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(bold=True, color=ACCENT)),
    )
    ws.freeze_panes = "C6"

    # ---- Cash flow --------------------------------------------------------
    ws = wb.create_sheet(s_cf)
    head(ws, 1, t["cf_head"], [12, 16, 16, 16, 26, 22])
    # Měsíc bez jediné transakce zůstane prázdný, ne nulový — teprve pak dávají
    # průměr i medián pod tabulkou smysl (obojí prázdné buňky přeskakuje).
    for m in range(MONTHS):
        r = m + 2
        col = get_column_letter(3 + m)
        empty = f'COUNTIF({q(s_tx)}$F$2:$F${ROWS_TX},$A{r})=0'
        ws.cell(row=r, column=1, value=f"={q(s_ov)}{col}${HR}").font = F_BODY
        for j, src in enumerate([r_inc, r_exp, r_cf, r_cfx]):
            c = ws.cell(row=r, column=2 + j, value=f'=IF({empty},"",{q(s_ov)}{col}{src})')
            c.number_format = t["money"]
            c.font = F_BODY
        ws.cell(row=r, column=6, value=f'=IF(D{r}="","",IF(D{r}<0,"{t["cf_neg"]}",""))').font = F_COMMENT

    last = MONTHS + 1
    ws.conditional_formatting.add(
        f"D2:E{last}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(bold=True, color=ACCENT)),
    )

    r = last + 2
    for label, fn in t["cf_sum"]:
        c = ws.cell(row=r, column=1, value=label)
        c.font = F_TOTAL
        for j in range(4):
            col = get_column_letter(2 + j)
            cc = ws.cell(row=r, column=2 + j, value=f"={fn}({col}2:{col}{last})")
            cc.number_format = t["money"]
            cc.font = F_TOTAL
        r += 1
    ws.cell(row=r, column=1, value=t["cf_negcount"]).font = F_TOTAL
    ws.cell(row=r, column=2, value=f'=COUNTIF(D2:D{last},"<0")').font = F_TOTAL
    note(ws, f"C{r}", t["cf_medhint"])

    ws.column_dimensions["H"].width = 52
    note(ws, "H1", t["cf_note"])
    ws.freeze_panes = "A2"

    # Pořadí listů: Transakce se plní jako první, číselník až za nimi.
    wb.move_sheet(s_cat, offset=1)
    wb.active = 0
    path = OUT / t["file"]
    wb.save(path)
    return path


# ============================================================ 2. jídelníček
# Zdroj: content/tipy/jidelnicek-a-nakupni-seznam-ai.mdx

MEALS = [
    # den, snídaně, dopolední svačina, oběd, odpolední svačina, večeře
    ("Pondělí", "Ovesná kaše s banánem", "Jablko", "Čočka na kyselo s vejcem", "Jogurt", "Kuřecí wrap se zeleninou"),
    ("Úterý", "Jogurt s müsli", "Mrkev", "Zbytek čočky", "Ořechy", "Těstoviny s rajčatovou omáčkou"),
    ("Středa", "Chleba s tvarohem", "Hruška", "Polévka a chleba", "Jogurt", "Kuřecí kari s rýží"),
    ("Čtvrtek", "Ovesná kaše s ořechy", "Banán", "Zbytek kari", "Jablko", "Zeleninový salát s vejcem"),
    ("Pátek", "Míchaná vejce s chlebem", "Mrkev", "Těstovinový salát", "Ořechy", "Pizza z lednice"),
    ("Sobota", "Lívance", "Jablko", "Pečené kuře s bramborem", "Jogurt", "Chlebíčky a polévka"),
    ("Neděle", "Vajíčka a pečivo", "Hruška", "Zbytek kuřete", "Ořechy", "Dýňová polévka"),
]

MEALS_EN = [
    ("Monday", "Porridge with banana", "Apple", "Lentil stew with an egg", "Yoghurt", "Chicken wrap with vegetables"),
    ("Tuesday", "Yoghurt with muesli", "Carrot", "Leftover lentils", "Nuts", "Pasta with tomato sauce"),
    ("Wednesday", "Bread with cottage cheese", "Pear", "Soup and bread", "Yoghurt", "Chicken curry with rice"),
    ("Thursday", "Porridge with nuts", "Banana", "Leftover curry", "Apple", "Vegetable salad with an egg"),
    ("Friday", "Scrambled eggs on toast", "Carrot", "Pasta salad", "Nuts", "Pizza from the fridge"),
    ("Saturday", "Pancakes", "Apple", "Roast chicken with potatoes", "Yoghurt", "Open sandwiches and soup"),
    ("Sunday", "Eggs and bread", "Pear", "Leftover chicken", "Nuts", "Pumpkin soup"),
]

# jídlo, surovina, množství, jednotka, kategorie(index), jen na jedno jídlo, mám doma
# Suroviny se schválně napříč dny opakují — právě jejich sečtení do jednoho
# řádku je to, co dělá nákupní seznam nákupním seznamem.
ING = [
    (0, "Ovesné vločky", 500, "g", 4, 0, 1),
    (0, "Banán", 3, "ks", 1, 0, 0),
    (0, "Čočka", 500, "g", 4, 0, 0),
    (0, "Vejce", 4, "ks", 2, 0, 0),
    (0, "Kuřecí prsa", 600, "g", 3, 0, 0),
    (0, "Tortilla", 6, "ks", 0, 0, 0),
    (0, "Paprika", 2, "ks", 1, 0, 0),
    (1, "Jogurt bílý", 2, "ks", 2, 0, 0),
    (1, "Müsli", 400, "g", 4, 0, 1),
    (1, "Mrkev", 5, "ks", 1, 0, 0),
    (1, "Těstoviny", 500, "g", 4, 0, 1),
    (1, "Rajčata drcená", 2, "ks", 4, 0, 0),
    (1, "Ořechy vlašské", 100, "g", 4, 0, 0),
    (2, "Chleba", 2, "ks", 0, 0, 0),
    (2, "Tvaroh", 2, "ks", 2, 0, 0),
    (2, "Hruška", 2, "ks", 1, 0, 0),
    (2, "Jogurt bílý", 2, "ks", 2, 0, 0),
    (2, "Rýže", 1, "kg", 4, 0, 1),
    (2, "Kari koření", 1, "bal", 4, 1, 0),
    (2, "Kokosové mléko", 1, "ks", 4, 1, 0),
    (3, "Ořechy vlašské", 100, "g", 4, 0, 0),
    (3, "Banán", 3, "ks", 1, 0, 0),
    (3, "Jablko", 3, "ks", 1, 0, 0),
    (3, "Salát ledový", 1, "ks", 1, 0, 0),
    (3, "Vejce", 4, "ks", 2, 0, 0),
    (4, "Vejce", 6, "ks", 2, 0, 0),
    (4, "Mrkev", 4, "ks", 1, 0, 0),
    (4, "Pizza mražená", 2, "ks", 5, 1, 0),
    (4, "Ořechy vlašské", 100, "g", 4, 0, 0),
    (5, "Mouka hladká", 1, "kg", 4, 0, 1),
    (5, "Mléko", 2, "l", 2, 0, 0),
    (5, "Jablko", 3, "ks", 1, 0, 0),
    (5, "Kuře celé", 1, "ks", 3, 0, 0),
    (5, "Brambory", 2, "kg", 1, 0, 0),
    (5, "Jogurt bílý", 2, "ks", 2, 0, 0),
    (5, "Chleba", 2, "ks", 0, 0, 0),
    (6, "Hruška", 2, "ks", 1, 0, 0),
    (6, "Ořechy vlašské", 100, "g", 4, 0, 0),
    (6, "Dýně hokaido", 1, "ks", 1, 1, 0),
    (6, "Smetana ke šlehání 33%", 1, "ks", 2, 0, 0),
    (6, "Prací prášek", 1, "bal", 6, 0, 0),
]

ING_EN = {
    "Ovesné vločky": "Rolled oats", "Banán": "Banana", "Čočka": "Lentils", "Vejce": "Eggs",
    "Kuřecí prsa": "Chicken breast", "Tortilla": "Tortillas", "Paprika": "Bell pepper",
    "Jogurt bílý": "Plain yoghurt", "Müsli": "Muesli", "Mrkev": "Carrots",
    "Těstoviny": "Pasta", "Rajčata drcená": "Chopped tomatoes", "Chleba": "Bread",
    "Tvaroh": "Cottage cheese", "Hruška": "Pears", "Rýže": "Rice",
    "Kari koření": "Curry powder", "Kokosové mléko": "Coconut milk",
    "Ořechy vlašské": "Walnuts", "Jablko": "Apples", "Salát ledový": "Iceberg lettuce",
    "Pizza mražená": "Frozen pizza", "Mouka hladká": "Plain flour", "Kuře celé": "Whole chicken",
    "Brambory": "Potatoes", "Mléko": "Milk", "Dýně hokaido": "Hokkaido pumpkin",
    "Smetana ke šlehání 33%": "Double cream 33%", "Prací prášek": "Laundry detergent",
}

UNITS_EN = {"g": "g", "ks": "pcs", "kg": "kg", "l": "l", "bal": "pack"}

MENU = {
    "cs": {
        "file": "jidelnicek-nakupni-seznam.xlsx",
        "sheets": ["Týden", "Suroviny", "Nákupní seznam"],
        "week_head": ["Den", "Snídaně", "Dopolední svačina", "Oběd", "Odpolední svačina", "Večeře", "Poznámka"],
        "week_note": (
            "Jídelníček na týden. Nechte si ho vygenerovat podle profilu domácnosti a pak "
            "přepište, co nesedí — plán, který se nedá změnit, nikdo nedodrží."
        ),
        "week_notes": [
            "", "Vaří se dvojitá dávka na středu", "", "", "Rychlý pátek", "Velké vaření", "Ze zbytků",
        ],
        "ing_head": ["Jídlo", "Surovina", "Množství", "Jednotka", "Kategorie", "Jen na jedno jídlo", "Mám doma"],
        "ing_note": (
            "Suroviny ke každému jídlu. Sloupec „Jen na jedno jídlo“ ukáže, co kupujete "
            "na jedno použití — bývá to přesně to, co zvedá cenu nákupu."
        ),
        "cats": ["pečivo", "ovoce a zelenina", "mléčné a vejce", "maso a ryby", "trvanlivé a koření", "mražené", "drogerie"],
        "yes": "ano",
        "no": "ne",
        "list_head": ["Kategorie", "Surovina", "Množství celkem", "Jednotka", "Duplicita", "Koupit", "Řádek do seznamu"],
        "list_note": (
            "Celý list se skládá vzorcem ze Surovin: stejné suroviny se sečtou, duplicitní "
            "řádky se označí × a co máte doma vypadne. Filtrujte sloupec „Koupit“ na "
            "„koupit“ a máte hotový seznam do obchodu."
        ),
        "buy": "koupit",
        "have": "mám doma",
        "dup": "—",
        "stock_title": "Zásoby — došlo?",
        "stock_head": ["Položka", "Došlo?"],
        "stock": ["sůl", "olej", "mouka", "cukr", "káva", "toaletní papír", "houbičky na nádobí"],
    },
    "en": {
        "file": "jidelnicek-nakupni-seznam-en.xlsx",
        "sheets": ["Week", "Ingredients", "Shopping list"],
        "week_head": ["Day", "Breakfast", "Morning snack", "Lunch", "Afternoon snack", "Dinner", "Note"],
        "week_note": (
            "The week's menu. Have it generated from your household profile, then overwrite "
            "whatever does not fit — a plan you cannot change is a plan nobody keeps."
        ),
        "week_notes": [
            "", "Double batch cooked for Wednesday", "", "", "Quick Friday", "Big cook", "Leftovers",
        ],
        "ing_head": ["Meal", "Ingredient", "Quantity", "Unit", "Aisle", "Single-use only", "Have at home"],
        "ing_note": (
            "Ingredients per meal. The “Single-use only” column shows what you are buying for "
            "one dish — usually exactly what pushes the shop over budget."
        ),
        "cats": ["bakery", "fruit and veg", "dairy and eggs", "meat and fish", "dry goods and spices", "frozen", "household"],
        "yes": "yes",
        "no": "no",
        "list_head": ["Aisle", "Ingredient", "Total quantity", "Unit", "Duplicate", "Buy", "Line for the list"],
        "list_note": (
            "This whole sheet is built by formula from Ingredients: identical items are added up, "
            "repeated rows are marked ×, and anything you already have drops out. Filter the "
            "“Buy” column to “buy” and you have your shopping list."
        ),
        "buy": "buy",
        "have": "have at home",
        "dup": "—",
        "stock_title": "Staples — run out?",
        "stock_head": ["Item", "Run out?"],
        "stock": ["salt", "oil", "flour", "sugar", "coffee", "toilet paper", "washing-up sponges"],
    },
}


def build_menu(loc: str) -> pathlib.Path:
    t = MENU[loc]
    s_week, s_ing, s_list = t["sheets"]
    meals = MEALS_EN if loc == "en" else MEALS

    wb = Workbook()

    # ---- Týden ------------------------------------------------------------
    ws = wb.active
    ws.title = s_week
    head(ws, 1, t["week_head"], [12, 26, 18, 26, 18, 26, 26])
    for i, row in enumerate(meals):
        r = i + 2
        ws.cell(row=r, column=1, value=row[0]).font = F_HEAD
        for j in range(1, 6):
            c = ws.cell(row=r, column=1 + j, value=row[j])
            c.font = F_BODY
            c.alignment = WRAP
        c = ws.cell(row=r, column=7, value=t["week_notes"][i])
        c.font = F_COMMENT
        c.alignment = WRAP
        ws.row_dimensions[r].height = 30
    ws.column_dimensions["I"].width = 50
    note(ws, "I1", t["week_note"])
    ws.freeze_panes = "B2"

    # ---- Suroviny ---------------------------------------------------------
    ws = wb.create_sheet(s_ing)
    head(ws, 1, t["ing_head"], [14, 26, 12, 11, 22, 18, 13])
    for i, (day, name, qty, unit, cat, single, home) in enumerate(ING):
        r = i + 2
        ws.cell(row=r, column=1, value=meals[day][0]).font = F_BODY
        ws.cell(row=r, column=2, value=ING_EN[name] if loc == "en" else name).font = F_BODY
        ws.cell(row=r, column=3, value=qty).font = F_BODY
        ws.cell(row=r, column=4, value=UNITS_EN[unit] if loc == "en" else unit).font = F_BODY
        ws.cell(row=r, column=5, value=t["cats"][cat]).font = F_BODY
        ws.cell(row=r, column=6, value=t["yes"] if single else t["no"]).font = F_BODY
        ws.cell(row=r, column=7, value=t["yes"] if home else t["no"]).font = F_BODY

    dv_aisle = DataValidation(type="list", formula1='"' + ",".join(t["cats"]) + '"', allow_blank=True)
    dv_yesno = DataValidation(type="list", formula1=f'"{t["yes"]},{t["no"]}"', allow_blank=True)
    ws.add_data_validation(dv_aisle)
    ws.add_data_validation(dv_yesno)
    dv_aisle.add(f"E2:E{ROWS_ING}")
    dv_yesno.add(f"F2:G{ROWS_ING}")

    ws.column_dimensions["I"].width = 50
    note(ws, "I1", t["ing_note"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ROWS_ING}"

    # ---- Nákupní seznam (celý vzorcem) ------------------------------------
    ws = wb.create_sheet(s_list)
    head(ws, 1, t["list_head"], [22, 26, 16, 11, 12, 14, 34])
    ing = q(s_ing)
    for r in range(2, ROWS_ING + 1):
        ws.cell(row=r, column=1, value=f'=IF({ing}$B{r}="","",{ing}$E{r})').font = F_BODY
        ws.cell(row=r, column=2, value=f'=IF({ing}$B{r}="","",{ing}$B{r})').font = F_BODY
        ws.cell(
            row=r, column=3,
            value=(
                f'=IF($B{r}="","",SUMIFS({ing}$C$2:$C${ROWS_ING},'
                f'{ing}$B$2:$B${ROWS_ING},$B{r},{ing}$D$2:$D${ROWS_ING},$D{r}))'
            ),
        ).font = F_BODY
        ws.cell(row=r, column=4, value=f'=IF({ing}$B{r}="","",{ing}$D{r})').font = F_BODY
        ws.cell(
            row=r, column=5,
            value=(
                f'=IF($B{r}="","",IF(COUNTIFS({ing}$B$2:$B{r},$B{r},'
                f'{ing}$D$2:$D{r},$D{r})>1,"×",""))'
            ),
        ).font = F_COMMENT
        ws.cell(
            row=r, column=6,
            value=(
                f'=IF($B{r}="","",IF($E{r}="×","{t["dup"]}",'
                f'IF({ing}$G{r}="{t["yes"]}","{t["have"]}","{t["buy"]}")))'
            ),
        ).font = F_BODY
        ws.cell(
            row=r, column=7,
            value=f'=IF($F{r}="{t["buy"]}",$B{r}&" — "&$C{r}&" "&$D{r},"")',
        ).font = F_BODY

    ws.cell(row=1, column=9, value=t["stock_title"]).font = F_SECTION
    for j, label in enumerate(t["stock_head"]):
        c = ws.cell(row=2, column=9 + j, value=label)
        c.font = F_HEAD
        c.fill = FILL_HEAD
        c.border = BORDER_HEAD
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["J"].width = 12
    for i, item in enumerate(t["stock"]):
        ws.cell(row=3 + i, column=9, value=item).font = F_BODY
        ws.cell(row=3 + i, column=10, value=t["no"]).font = F_BODY
    dv_stock = DataValidation(type="list", formula1=f'"{t["yes"]},{t["no"]}"', allow_blank=True)
    ws.add_data_validation(dv_stock)
    dv_stock.add(f"J3:J{2 + len(t['stock'])}")

    ws.column_dimensions["L"].width = 52
    note(ws, "L1", t["list_note"])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ROWS_ING}"

    wb.active = 0
    path = OUT / t["file"]
    wb.save(path)
    return path


# ========================================================== 3. hodinová sazba
# Zdroj: components/RateCalculator.tsx + content/prirucka/cas-jako-zbozi.mdx

RATE = {
    "cs": {
        "file": "kalkulace-hodinove-sazby.xlsx",
        "sheet": "Kalkulace",
        "title": "Kalkulace hodinové sazby",
        "note": (
            "Metodika je stejná jako v kapitole „Čas jako zboží“ a v kalkulačce na webu: "
            "sazba není odměna za hodinu, ale podíl — kolik potřebujete za rok získat "
            "dělíte hodinami, které opravdu prodáte. Přepisujte jen modrá pole ve Vstupech, "
            "zbytek se počítá sám."
        ),
        "sec_in": "Vstupy — vyplňte podle sebe",
        "sec_calc": "Výpočet — počítá se sám",
        "sec_out": "Výsledek",
        "head_in": ["Položka", "Hodnota", "Jednotka", "Komentář"],
        "head_calc": ["Krok", "Hodnota", "Jednotka", "Jak se to počítá"],
        "money": MONEY,
        "u_year": "Kč / rok",
        "u_month": "Kč / měsíc",
        "u_weeks": "týdnů",
        "u_hours": "hodin",
        "u_pct": "%",
        "u_rate": "Kč / h",
        "u_czk": "Kč",
        "u_projh": "fakturovatelných hodin",
        "inputs": [
            ("Kolik chci ročně vydělat čistého", 600000, MONEY, "u_year",
             "Co má reálně zbýt vám: nájem, jídlo, rodina, splátky. Vezměte to z bankovního výpisu, ne z hlavy — odhad bývá nižší než skutečnost."),
            ("Měsíční provozní náklady", 8000, MONEY, "u_month",
             "Technika, software, účetní, pojištění, telefon, doprava, vzdělávání, prostor."),
            ("Pracovních týdnů v roce", 44, "0", "u_weeks",
             "52 minus dovolená, státní svátky, nemoc a rezerva na věci, které se naplánovat nedají."),
            ("Hodin práce týdně", 40, "0", "u_hours",
             "Kolik hodin týdně u práce reálně sedíte — včetně těch, které se nikomu neúčtují."),
            ("Podíl fakturovatelných hodin", 0.60, PCT, "u_pct",
             "Akvizice, administrativa, učení a prostoje se neplatí. Odhad tady nestačí — změřte si jeden poctivý měsíc."),
            ("Rezerva na daně a odvody", 0.30, PCT, "u_pct",
             "Podíl z každé faktury, který si hned odkládáte stranou. Sazbu si doplňte podle svého režimu — tahle tabulka ji za vás neurčuje."),
            ("Rezerva na výpadky a investice", 0.15, PCT, "u_pct",
             "Finanční polštář, nové vybavení, kurzy, důchod. Co se nedostane do sazby, to nikdy nevznikne."),
            ("Modelová zakázka", 10, "0", "u_projh",
             "Kolik hodin má typická zakázka. Spočítá se, kolik musí nést, aby se vyplatila."),
        ],
        "calc": [
            ("Roční provozní náklady", "={c}*12", MONEY, "u_year", "Měsíční náklady krát dvanáct. Roční součet, ne měsíční pocit."),
            ("Roční potřeba (živobytí + náklady)", "={net}+{ac}", MONEY, "u_year", "Kolik peněz musí za rok projít, aby vám zbylo, co jste si mysleli."),
            ("Obrat, který musíte vyfakturovat", "={need}/(1-{tax})", MONEY, "u_year", "Daně a odvody jdou do ceny nahoru, ne jako překvapení v březnu."),
            ("Prodejné hodiny za rok", "={w}*{h}*{b}", "0", "u_hours", "Týdny krát hodiny krát fakturovatelný podíl. Bývá to zhruba polovina prvního odhadu — to není chyba výpočtu, to je realita."),
            ("Minimální hodinová sazba", '=IF({bh}>0,{rev}/{bh},"")', MONEY, "u_rate", "Obrat dělený prodejnými hodinami. Spodní hranice, ne cílová cena — kolik se dá dostat, řekne trh."),
            ("Doporučená sazba s rezervou", "={min}*(1+{m})", MONEY, "u_rate", "Sazba, která navíc zaplatí hluchá období, nové vybavení a čas investovaný do sebe."),
        ],
        "out": [
            ("Minimální hodinová sazba", "={min}", MONEY, "u_rate", "Pod tímhle číslem vás práce stojí víc, než vynese."),
            ("Doporučená sazba s rezervou", "={rec}", MONEY, "u_rate", "Číslo, se kterým jdete do nabídky."),
            ("Modelová zakázka — minimum", "={min}*{p}", MONEY, "u_czk", "Zakázka pod touhle částkou se dotuje vaším volným časem."),
            ("Modelová zakázka — s rezervou", "={rec}*{p}", MONEY, "u_czk", "Stejná zakázka počítaná doporučenou sazbou."),
        ],
        "footer": (
            "Sazba se sama neaktualizuje. Přepočítejte kalkulaci jednou ročně, ve stejný "
            "termín jako daňové přiznání — je to nejlevnější ochrana proti pomalému chudnutí."
        ),
    },
    "en": {
        "file": "kalkulace-hodinove-sazby-en.xlsx",
        "sheet": "Rate calculation",
        "title": "Hourly rate calculation",
        "note": (
            "The method is the one from the chapter “Time as a product” and the calculator on "
            "the site: a rate is not a reward for an hour, it is a ratio — what you need to "
            "earn in a year divided by the hours you actually sell. Overwrite the Inputs only; "
            "the rest calculates itself."
        ),
        "sec_in": "Inputs — fill in your own figures",
        "sec_calc": "Calculation — done for you",
        "sec_out": "Result",
        "head_in": ["Item", "Value", "Unit", "Comment"],
        "head_calc": ["Step", "Value", "Unit", "How it is calculated"],
        "money": MONEY_EN,
        "u_year": "CZK / year",
        "u_month": "CZK / month",
        "u_weeks": "weeks",
        "u_hours": "hours",
        "u_pct": "%",
        "u_rate": "CZK / h",
        "u_czk": "CZK",
        "u_projh": "billable hours",
        "inputs": [
            ("What I want to take home a year", 600000, MONEY_EN, "u_year",
             "What has to actually stay with you: rent, food, family, repayments. Take it from your bank statement, not from memory — the guess is always lower than reality."),
            ("Monthly running costs", 8000, MONEY_EN, "u_month",
             "Gear, software, accountant, insurance, phone, travel, training, workspace."),
            ("Working weeks a year", 44, "0", "u_weeks",
             "52 minus holidays, public holidays, sick days and a buffer for what cannot be planned."),
            ("Hours worked a week", 40, "0", "u_hours",
             "How many hours you really spend on work — including the ones nobody gets billed for."),
            ("Share of billable hours", 0.60, PCT, "u_pct",
             "Sales, admin, learning and downtime are unpaid. A guess is not enough here — measure one honest month."),
            ("Reserve for tax and levies", 0.30, PCT, "u_pct",
             "The share of every invoice you set aside straight away. Put in the rate that fits your own situation — this sheet does not decide it for you."),
            ("Reserve for gaps and investment", 0.15, PCT, "u_pct",
             "A financial cushion, new equipment, courses, retirement. Whatever does not make it into the rate will never exist."),
            ("Model project", 10, "0", "u_projh",
             "How many hours a typical project takes. The sheet works out what it has to bring in."),
        ],
        "calc": [
            ("Annual running costs", "={c}*12", MONEY_EN, "u_year", "Monthly costs times twelve. An annual total, not a monthly feeling."),
            ("Annual need (living + costs)", "={net}+{ac}", MONEY_EN, "u_year", "How much money has to pass through the year for you to keep what you planned."),
            ("Revenue you have to invoice", "={need}/(1-{tax})", MONEY_EN, "u_year", "Tax and levies go into the price up front, not as a surprise later."),
            ("Sellable hours a year", "={w}*{h}*{b}", "0", "u_hours", "Weeks times hours times the billable share. It usually comes out around half the first guess — that is not an error, that is reality."),
            ("Minimum hourly rate", '=IF({bh}>0,{rev}/{bh},"")', MONEY_EN, "u_rate", "Revenue divided by sellable hours. A floor, not a target price — what you can actually get is up to the market."),
            ("Recommended rate with a reserve", "={min}*(1+{m})", MONEY_EN, "u_rate", "The rate that also pays for quiet periods, new equipment and time invested in yourself."),
        ],
        "out": [
            ("Minimum hourly rate", "={min}", MONEY_EN, "u_rate", "Below this number the work costs you more than it brings in."),
            ("Recommended rate with a reserve", "={rec}", MONEY_EN, "u_rate", "The number you take into a quote."),
            ("Model project — minimum", "={min}*{p}", MONEY_EN, "u_czk", "A project below this is subsidised by your own free time."),
            ("Model project — with a reserve", "={rec}*{p}", MONEY_EN, "u_czk", "The same project at the recommended rate."),
        ],
        "footer": (
            "A rate does not update itself. Recalculate once a year, on the same date as your "
            "tax return — it is the cheapest protection against getting slowly poorer."
        ),
    },
}


def build_rate(loc: str) -> pathlib.Path:
    t = RATE[loc]
    wb = Workbook()
    ws = wb.active
    ws.title = t["sheet"]
    for col, w in zip("ABCD", (38, 16, 22, 78)):
        ws.column_dimensions[col].width = w

    ws["A1"] = t["title"]
    ws["A1"].font = F_TITLE
    ws["A2"] = t["note"]
    ws["A2"].font = F_NOTE
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 42

    r = 4
    ws.cell(row=r, column=1, value=t["sec_in"]).font = F_SECTION
    r += 1
    head(ws, r, t["head_in"])
    r += 1
    first_in = r
    for label, value, fmt, unit, comment in t["inputs"]:
        ws.cell(row=r, column=1, value=label).font = F_BODY
        c = ws.cell(row=r, column=2, value=value)
        c.number_format = fmt
        c.font = F_TOTAL
        ws.cell(row=r, column=3, value=t[unit]).font = F_COMMENT
        cc = ws.cell(row=r, column=4, value=comment)
        cc.font = F_COMMENT
        cc.alignment = WRAP
        r += 1

    B = {k: f"$B${first_in + i}" for i, k in enumerate(
        ["net", "c", "w", "h", "b", "tax", "m", "p"]
    )}

    r += 1
    ws.cell(row=r, column=1, value=t["sec_calc"]).font = F_SECTION
    r += 1
    head(ws, r, t["head_calc"])
    r += 1
    first_calc = r
    keys = ["ac", "need", "rev", "bh", "min", "rec"]
    refs = dict(B)
    for i, k in enumerate(keys):
        refs[k] = f"$B${first_calc + i}"
    for i, (label, formula, fmt, unit, comment) in enumerate(t["calc"]):
        ws.cell(row=r, column=1, value=label).font = F_BODY
        c = ws.cell(row=r, column=2, value=formula.format(**refs))
        c.number_format = fmt
        c.font = F_TOTAL
        ws.cell(row=r, column=3, value=t[unit]).font = F_COMMENT
        cc = ws.cell(row=r, column=4, value=comment)
        cc.font = F_COMMENT
        cc.alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=t["sec_out"]).font = F_SECTION
    r += 1
    for label, formula, fmt, unit, comment in t["out"]:
        c = ws.cell(row=r, column=1, value=label)
        c.font = F_TOTAL
        c.border = BORDER_TOP
        c = ws.cell(row=r, column=2, value=formula.format(**refs))
        c.number_format = fmt
        c.font = F_RESULT
        c.border = BORDER_TOP
        c = ws.cell(row=r, column=3, value=t[unit])
        c.font = F_COMMENT
        c.border = BORDER_TOP
        cc = ws.cell(row=r, column=4, value=comment)
        cc.font = F_COMMENT
        cc.alignment = WRAP
        cc.border = BORDER_TOP
        r += 1

    r += 1
    c = ws.cell(row=r, column=1, value=t["footer"])
    c.font = F_NOTE
    c.alignment = WRAP

    path = OUT / t["file"]
    wb.save(path)
    return path


# ================================================================== 4-6. CSV

def write_csv(name: str, rows: list[list[str]]) -> pathlib.Path:
    path = OUT / name
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh, delimiter=",", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerows(rows)
    return path


# --- 4. Vizitky do CRM -------------------------------------------------------
# Zdroj: content/tipy/vizitky-do-crm-claude-cowork.mdx, Fáze 2 (mapování hlavičky).
# Hlavička bez diakritiky a bez mezer — tak ji čekají importy běžných CRM.

CRM = {
    "cs": {
        "file": "vizitky-do-crm.csv",
        "header": [
            "jmeno", "prijmeni", "pozice", "firma", "email", "telefon", "web",
            "mesto", "zeme", "linkedin", "zdroj_kontaktu", "datum_kontaktu",
            "stitky", "poznamka_z_vizitky", "dalsi_krok", "termin_dalsiho_kroku", "vlastnik",
        ],
        "row": [
            "Petra", "Havlíčková", "vedoucí zákaznické podpory", "Vltavín Design s.r.o.",
            "petra.havlickova@example.com", "+420 601 000 000", "https://www.example.com",
            "Brno", "CZ", "https://www.linkedin.com/in/example",
            "Konference Firma a AI 2026", "2026-05-12",
            "AI, zákaznická podpora",
            "Zdroj: konference Firma a AI 2026, 12. 5. 2026. Poznámka z vizitky: řeší tříděné odpovědi na e-maily, ozvat se po prázdninách.",
            "poslat případovou studii", "2026-09-02", "jan.novak@example.com",
        ],
    },
    "en": {
        "file": "vizitky-do-crm-en.csv",
        "header": [
            "first_name", "last_name", "job_title", "company", "email", "phone", "website",
            "city", "country", "linkedin", "lead_source", "contact_date",
            "tags", "business_card_note", "next_step", "next_step_due", "owner",
        ],
        "row": [
            "Petra", "Havlickova", "head of customer support", "Vltavin Design Ltd.",
            "petra.havlickova@example.com", "+420 601 000 000", "https://www.example.com",
            "Brno", "CZ", "https://www.linkedin.com/in/example",
            "Business and AI 2026 conference", "2026-05-12",
            "AI, customer support",
            "Source: Business and AI 2026 conference, 12 May 2026. Card note: working on triaging inbound email, follow up after the summer.",
            "send the case study", "2026-09-02", "jan.novak@example.com",
        ],
    },
}


def build_crm(loc: str) -> pathlib.Path:
    t = CRM[loc]
    return write_csv(t["file"], [t["header"], t["row"]])


# --- 5. Rodinný kalendář -----------------------------------------------------
# Zdroj: content/tipy/ai-v-rodine-vychova-hry-pohadky.mdx (rodinný provoz)
#      + content/prirucka/domacnost-jako-provoz.mdx (agendy a rytmus).
# Bez jmen dětí — návod záměrně doporučuje „starší dítě“ / „mladší dítě“.

FAMILY_ROWS = [
    ("Plavání", "Kroužek", "Úterý", "16:00", "17:00", "starší dítě", "bazén Na Lukách", "rodič A", "týdně", "plavky, ručník, čepice", "ze školy pěšky 15 minut"),
    ("Keramika", "Kroužek", "Středa", "15:30", "17:00", "mladší dítě", "DDM", "rodič B", "týdně", "staré tričko", ""),
    ("Fotbal trénink", "Kroužek", "Čtvrtek", "17:00", "18:30", "starší dítě", "hřiště u školy", "rodič A", "týdně", "kopačky, chrániče", "kolize s keramikou, pokud se protáhne"),
    ("Zubař kontrola", "Zdraví", "Pondělí", "09:30", "10:00", "oba", "ordinace Kollárova", "rodič B", "jednorázově", "kartička pojišťovny", "termín 2026-09-14"),
    ("Třídní schůzky", "Škola", "Čtvrtek", "17:30", "19:00", "rodič", "škola, třída 3.B", "rodič A", "jednorázově", "", "termín 2026-09-24"),
    ("Velký nákup", "Domácnost", "Sobota", "09:00", "10:30", "rodina", "obchod", "rodič B", "týdně", "nákupní seznam", "vychází z jídelníčku na týden"),
    ("Praní a žehlení", "Domácnost", "Neděle", "18:00", "19:00", "rodina", "doma", "rodič A", "týdně", "", "agenda, ne jednorázový úkol"),
    ("Kontrola zásob a drogerie", "Domácnost", "Pátek", "18:00", "18:20", "rodina", "doma", "rodič B", "týdně", "", "co dochází, jde rovnou do seznamu"),
    ("Úklid společných prostor", "Domácnost", "Sobota", "11:00", "12:00", "rodina", "doma", "rodina", "týdně", "", ""),
    ("Papíry a platby", "Domácnost", "Neděle", "19:00", "19:30", "rodič", "doma", "rodič A", "měsíčně", "", "spolu s měsíční uzávěrkou rozpočtu"),
    ("Přezutí kol a servis auta", "Domácnost", "Sobota", "", "", "rodič", "servis", "rodič B", "sezónně", "", "říjen a duben"),
    ("Předání dětí", "Předávání", "Neděle", "18:00", "18:15", "oba", "u druhého rodiče", "rodič A", "týdně", "školní batohy, plavky, nabíječka, léky", "shrnutí druhému rodiči piš věcně, bez hodnocení"),
]

FAMILY_ROWS_EN = [
    ("Swimming", "Club", "Tuesday", "16:00", "17:00", "older child", "Na Lukach pool", "parent A", "weekly", "swimsuit, towel, cap", "15-minute walk from school"),
    ("Pottery", "Club", "Wednesday", "15:30", "17:00", "younger child", "community centre", "parent B", "weekly", "old T-shirt", ""),
    ("Football practice", "Club", "Thursday", "17:00", "18:30", "older child", "pitch by the school", "parent A", "weekly", "boots, shin pads", "clashes with pottery if it overruns"),
    ("Dentist check-up", "Health", "Monday", "09:30", "10:00", "both", "Kollarova surgery", "parent B", "one-off", "insurance card", "date 2026-09-14"),
    ("Parents' evening", "School", "Thursday", "17:30", "19:00", "parent", "school, class 3B", "parent A", "one-off", "", "date 2026-09-24"),
    ("Big shop", "Household", "Saturday", "09:00", "10:30", "family", "supermarket", "parent B", "weekly", "shopping list", "comes out of the weekly menu"),
    ("Laundry and ironing", "Household", "Sunday", "18:00", "19:00", "family", "home", "parent A", "weekly", "", "an agenda, not a one-off task"),
    ("Stock and household supplies check", "Household", "Friday", "18:00", "18:20", "family", "home", "parent B", "weekly", "", "whatever runs low goes straight on the list"),
    ("Cleaning the shared rooms", "Household", "Saturday", "11:00", "12:00", "family", "home", "family", "weekly", "", ""),
    ("Paperwork and payments", "Household", "Sunday", "19:00", "19:30", "parent", "home", "parent A", "monthly", "", "together with the monthly budget close"),
    ("Tyre change and car service", "Household", "Saturday", "", "", "parent", "garage", "parent B", "seasonally", "", "October and April"),
    ("Handover", "Handover", "Sunday", "18:00", "18:15", "both", "at the other parent's", "parent A", "weekly", "school bags, swimsuits, charger, medicines", "keep the summary for the other parent factual, no judgement"),
]

FAMILY = {
    "cs": {
        "file": "rodinny-kalendar.csv",
        "header": ["Název", "Typ", "Den", "Čas od", "Čas do", "Kdo", "Místo",
                   "Kdo doprovází", "Rytmus", "Vzít s sebou", "Poznámka"],
        "rows": FAMILY_ROWS,
    },
    "en": {
        "file": "rodinny-kalendar-en.csv",
        "header": ["Name", "Type", "Day", "Time from", "Time to", "Who", "Place",
                   "Who takes them", "Rhythm", "What to bring", "Note"],
        "rows": FAMILY_ROWS_EN,
    },
}


def build_family(loc: str) -> pathlib.Path:
    t = FAMILY[loc]
    return write_csv(t["file"], [t["header"]] + [list(r) for r in t["rows"]])


# --- 6. Projekty a úkoly -----------------------------------------------------
# Zdroj: content/tipy/ai-ve-firme-kompletni-pruvodce.mdx, část 5 „Krok 2:
# struktura v Notionu“. Projekty i úkoly v jedné databázi, sloupec „Typ záznamu“
# je rozlišuje; vazba na projekt je text, po importu z něj v Notionu uděláte relaci.
# Meduna s.r.o. je vymyšlená firma z toho návodu.

WORK_ROWS = [
    ("Nový web a ceník", "Projekt", "", "Jana Dvořáková", "Běží", "2026-10-31", "P1", "Marketing", "2026-08-11", "Porada vedení 2026-08-10", "Původní web má čtyři roky a nejde do něj psát."),
    ("Automatizace faktur od dodavatelů", "Projekt", "", "Petr Málek", "Riziko", "2026-09-30", "P1", "Provoz", "2026-07-29", "Porada provozu 2026-07-28", "Tři týdny ticho — hlídač skluzu tenhle projekt vytáhl první."),
    ("Firemní AI směrnice", "Projekt", "", "Jana Dvořáková", "Čeká na externí", "2026-09-15", "P2", "Vedení", "2026-08-05", "Porada vedení 2026-08-03", "Čeká na vyjádření právničky ke klasifikaci dat."),
    ("Onboarding nových lidí", "Projekt", "", "Lucie Hrubá", "Nový", "2026-12-15", "P3", "HR", "2026-08-12", "Porada HR 2026-08-12", "Zatím jen záměr, rozsah není domluvený."),
    ("Sepsat strukturu ceníku", "Úkol", "Nový web a ceník", "Jana Dvořáková", "Běží", "2026-08-28", "P1", "Marketing", "2026-08-11", "Porada vedení 2026-08-10", ""),
    ("Vybrat dodavatele fotek", "Úkol", "Nový web a ceník", "Tomáš Beneš", "Čeká", "2026-09-04", "P2", "Marketing", "2026-08-10", "Porada vedení 2026-08-10", ""),
    ("Otestovat OCR na 50 fakturách", "Úkol", "Automatizace faktur od dodavatelů", "Petr Málek", "Blokováno", "2026-08-22", "P1", "Provoz", "2026-07-29", "Porada provozu 2026-07-28", "Blokuje přístup do účetního systému."),
    ("Popsat schvalovací kolečko faktur", "Úkol", "Automatizace faktur od dodavatelů", "Petr Málek", "Čeká", "2026-09-05", "P2", "Provoz", "2026-07-28", "Porada provozu 2026-07-28", ""),
    ("Připomínkovat vzor směrnice", "Úkol", "Firemní AI směrnice", "Jana Dvořáková", "Hotovo", "2026-08-04", "P2", "Vedení", "2026-08-05", "Porada vedení 2026-08-03", ""),
    ("Doplnit klasifikaci dat do směrnice", "Úkol", "Firemní AI směrnice", "externí právnička", "Čeká", "2026-09-08", "P1", "Vedení", "2026-08-05", "Porada vedení 2026-08-03", "Externí termín — hlídat, ne tlačit."),
    ("Sepsat půldenní školení pro tým", "Úkol", "Firemní AI směrnice", "Lucie Hrubá", "Čeká", "2026-09-19", "P3", "HR", "2026-08-12", "Porada HR 2026-08-12", ""),
    ("Zjistit, co lidé používají dnes", "Úkol", "", "Lucie Hrubá", "Běží", "2026-08-26", "P2", "HR", "2026-08-12", "Porada HR 2026-08-12", "NOVÝ PROJEKT? Vypadá to na inventuru shadow AI."),
]

WORK_ROWS_EN = [
    ("New website and price list", "Project", "", "Jana Dvorakova", "Running", "2026-10-31", "P1", "Marketing", "2026-08-11", "Leadership meeting 2026-08-10", "The current site is four years old and cannot be edited."),
    ("Supplier invoice automation", "Project", "", "Petr Malek", "At risk", "2026-09-30", "P1", "Operations", "2026-07-29", "Operations meeting 2026-07-28", "Three weeks of silence — the drift watcher surfaced this one first."),
    ("Company AI policy", "Project", "", "Jana Dvorakova", "Waiting on external", "2026-09-15", "P2", "Leadership", "2026-08-05", "Leadership meeting 2026-08-03", "Waiting for the lawyer's take on data classification."),
    ("Onboarding for new hires", "Project", "", "Lucie Hruba", "New", "2026-12-15", "P3", "HR", "2026-08-12", "HR meeting 2026-08-12", "An intention so far, the scope is not agreed."),
    ("Draft the price list structure", "Task", "New website and price list", "Jana Dvorakova", "Running", "2026-08-28", "P1", "Marketing", "2026-08-11", "Leadership meeting 2026-08-10", ""),
    ("Pick a photographer", "Task", "New website and price list", "Tomas Benes", "Waiting", "2026-09-04", "P2", "Marketing", "2026-08-10", "Leadership meeting 2026-08-10", ""),
    ("Test OCR on 50 invoices", "Task", "Supplier invoice automation", "Petr Malek", "Blocked", "2026-08-22", "P1", "Operations", "2026-07-29", "Operations meeting 2026-07-28", "Blocked on access to the accounting system."),
    ("Write up the invoice approval loop", "Task", "Supplier invoice automation", "Petr Malek", "Waiting", "2026-09-05", "P2", "Operations", "2026-07-28", "Operations meeting 2026-07-28", ""),
    ("Review the draft policy", "Task", "Company AI policy", "Jana Dvorakova", "Done", "2026-08-04", "P2", "Leadership", "2026-08-05", "Leadership meeting 2026-08-03", ""),
    ("Add data classification to the policy", "Task", "Company AI policy", "external lawyer", "Waiting", "2026-09-08", "P1", "Leadership", "2026-08-05", "Leadership meeting 2026-08-03", "An external deadline — watch it, do not push it."),
    ("Prepare the half-day team training", "Task", "Company AI policy", "Lucie Hruba", "Waiting", "2026-09-19", "P3", "HR", "2026-08-12", "HR meeting 2026-08-12", ""),
    ("Find out what people already use", "Task", "", "Lucie Hruba", "Running", "2026-08-26", "P2", "HR", "2026-08-12", "HR meeting 2026-08-12", "NEW PROJECT? This looks like a shadow AI inventory."),
]

WORK = {
    "cs": {
        "file": "projekty-a-ukoly.csv",
        "header": ["Název", "Typ záznamu", "Projekt", "Vlastník / řešitel", "Stav",
                   "Termín", "Priorita", "Oblast", "Poslední aktivita", "Zdroj", "Poznámka"],
        "rows": WORK_ROWS,
    },
    "en": {
        "file": "projekty-a-ukoly-en.csv",
        "header": ["Name", "Record type", "Project", "Owner / assignee", "Status",
                   "Due", "Priority", "Area", "Last activity", "Source", "Note"],
        "rows": WORK_ROWS_EN,
    },
}


def build_work(loc: str) -> pathlib.Path:
    t = WORK[loc]
    return write_csv(t["file"], [t["header"]] + [list(r) for r in t["rows"]])


# =================================================================== spuštění

BUILDERS = [build_budget, build_menu, build_rate, build_crm, build_family, build_work]


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    made: list[pathlib.Path] = []
    for build in BUILDERS:
        for loc in ("cs", "en"):
            made.append(build(loc))
    for path in sorted(made):
        size = os.path.getsize(path)
        print(f"{path.relative_to(ROOT)}  {size / 1024:.1f} kB")


if __name__ == "__main__":
    main()
